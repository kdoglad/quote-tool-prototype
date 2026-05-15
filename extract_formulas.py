import openpyxl
import csv
import sys
import re

excel_file = "XXkWp_Project Name_PCC_INTERNAL_V22.9 - DO NOT OPEN MAKE A COPY.xlsm"
csv_file = "formulas_extracted.csv"

# Regex to match simple single-cell reference formulas
simple_cell_ref_pattern = re.compile(
    r'^\s*=\s*\+?\s*\(?\s*(?:\'[^\']+\'!|[A-Za-z0-9_]+!)?\$?[A-Za-z]+\$?[0-9]+\s*\)?\s*$',
    re.IGNORECASE
)

def is_valid_label(val):
    if not isinstance(val, str):
        return False
    s = val.strip()
    if len(s) < 2:
        return False
    if s.startswith('='):
        return False
    # Check if purely numeric
    try:
        float(s)
        return False
    except ValueError:
        pass
    return True

def clean_label_str(s):
    # Remove excessive whitespace/newlines
    return re.sub(r'\s+', ' ', s).strip()

def get_sheet_by_name(wb, sheet_str, default_sheet_name):
    if sheet_str:
        # Strip trailing ! and quotes
        s_name = sheet_str.rstrip('!').strip("'").strip()
        if s_name in wb.sheetnames:
            return wb[s_name], s_name
    return wb[default_sheet_name], default_sheet_name

def resolve_cell_label(wb, current_sheet_name, sheet_str, ref_col_str, ref_row_int, formula_row_int):
    target_sheet, s_name = get_sheet_by_name(wb, sheet_str, current_sheet_name)
    try:
        col_idx = openpyxl.utils.column_index_from_string(ref_col_str)
    except Exception:
        return f"{ref_col_str}{ref_row_int}"
        
    col_header = None
    if 1 <= col_idx <= target_sheet.max_column and ref_row_int > 1:
        start_r = min(ref_row_int - 1, target_sheet.max_row)
        end_r = max(1, ref_row_int - 50)
        for r in range(start_r, end_r - 1, -1):
            val = target_sheet.cell(row=r, column=col_idx).value
            if is_valid_label(val):
                col_header = clean_label_str(val)
                break
            
    row_label = None
    if 1 <= ref_row_int <= target_sheet.max_row:
        # Check standard label columns (1 and 2)
        for c in range(min(col_idx - 1, 2), 0, -1):
            val = target_sheet.cell(row=ref_row_int, column=c).value
            if is_valid_label(val):
                row_label = clean_label_str(val)
                break
            
    # Include sheet prefix if referencing a different sheet
    prefix = f"{s_name} " if s_name != current_sheet_name else ""
    
    # Strategy selection
    if ref_row_int == formula_row_int and s_name == current_sheet_name:
        # Same row reference in same sheet -> prefer column header
        if col_header:
            return f"{prefix}{col_header}"
        if row_label:
            return f"{prefix}{row_label}"
    else:
        # Different row or different sheet -> prefer row label
        if row_label:
            return f"{prefix}{row_label}"
        if col_header:
            return f"{prefix}{col_header}"
            
    return f"{sheet_str or ''}{ref_col_str}{ref_row_int}"

def resolve_lhs_name(sheet, cell_coord, row_int, col_int):
    if 1 <= row_int <= sheet.max_row:
        # Check standard label columns (1 and 2)
        for c in range(min(col_int - 1, 2), 0, -1):
            val = sheet.cell(row=row_int, column=c).value
            if is_valid_label(val):
                return clean_label_str(val)
                
    if 1 <= col_int <= sheet.max_column and row_int > 1:
        start_r = min(row_int - 1, sheet.max_row)
        end_r = max(1, row_int - 50)
        for r in range(start_r, end_r - 1, -1):
            val = sheet.cell(row=r, column=col_int).value
            if is_valid_label(val):
                return clean_label_str(val)
                
    return cell_coord

# Regex patterns for parsing formulas
# 1. SUM(range) pattern
sum_range_regex = re.compile(
    r'SUM\(\s*(?P<sheet>(?:\'[^\']+\'|[A-Za-z0-9_]+)\!)?\$?(?P<c1>[A-Za-z]+)\$?(?P<r1>[0-9]+):\$?(?P<c2>[A-Za-z]+)\$?(?P<r2>[0-9]+)\s*\)',
    re.IGNORECASE
)

# 2. General range pattern (outside already resolved tokens)
general_range_regex = re.compile(
    r'(\x01[^\x02]*\x02)|(?:(?P<sheet>(?:\'[^\']+\'|[A-Za-z0-9_]+)\!)?\$?(?P<c1>[A-Za-z]+)\$?(?P<r1>[0-9]+):\$?(?P<c2>[A-Za-z]+)\$?(?P<r2>[0-9]+))'
)

# 3. Single cell reference pattern (outside already resolved tokens)
single_cell_regex = re.compile(
    r'(\x01[^\x02]*\x02)|(?:(?P<sheet>(?:\'[^\']+\'|[A-Za-z0-9_]+)\!)?\$?(?P<col>[A-Za-z]+)\$?(?P<row>[0-9]+))'
)

def translate_formula_to_words(wb, sheet_name, cell, formula_str):
    formula_row = cell.row
    formula_col = cell.column
    
    # Resolve LHS name
    lhs_name = resolve_lhs_name(wb[sheet_name], cell.coordinate, formula_row, formula_col)
    
    # Strip leading '='
    expr = formula_str.lstrip('=').strip()
    
    # Pass 1: Substitute SUM(range) calls
    def repl_sum_range(match):
        sheet_str = match.group('sheet')
        c1, r1 = match.group('c1'), int(match.group('r1'))
        c2, r2 = match.group('c2'), int(match.group('r2'))
        
        target_sheet, s_name = get_sheet_by_name(wb, sheet_str, sheet_name)
        try:
            col1 = openpyxl.utils.column_index_from_string(c1)
            col2 = openpyxl.utils.column_index_from_string(c2)
        except Exception:
            return match.group(0)
            
        if r1 <= r2 and col1 <= col2:
            total_cells = (r2 - r1 + 1) * (col2 - col1 + 1)
            if total_cells <= 12:
                labels = []
                for r in range(r1, r2 + 1):
                    for c in range(col1, col2 + 1):
                        c_letter = openpyxl.utils.get_column_letter(c)
                        lbl = resolve_cell_label(wb, sheet_name, sheet_str, c_letter, r, formula_row)
                        labels.append(lbl)
                joined = " + ".join(labels)
                # If the entire expression was exactly this SUM(), no outer parens needed
                if match.group(0).strip() == expr:
                    return f"\x01{joined}\x02"
                return f"\x01({joined})\x02"
                
        # Fallback for large SUM ranges
        start_lbl = resolve_cell_label(wb, sheet_name, sheet_str, c1, r1, formula_row)
        end_lbl = resolve_cell_label(wb, sheet_name, sheet_str, c2, r2, formula_row)
        return f"\x01Sum of {start_lbl} through {end_lbl}\x02"
        
    expr = sum_range_regex.sub(repl_sum_range, expr)
    
    # Pass 2: Substitute general ranges
    def repl_general_range(match):
        if match.group(1):
            return match.group(1)
        sheet_str = match.group('sheet')
        c1, r1 = match.group('c1'), int(match.group('r1'))
        c2, r2 = match.group('c2'), int(match.group('r2'))
        start_lbl = resolve_cell_label(wb, sheet_name, sheet_str, c1, r1, formula_row)
        end_lbl = resolve_cell_label(wb, sheet_name, sheet_str, c2, r2, formula_row)
        return f"\x01range {start_lbl} to {end_lbl}\x02"
        
    expr = general_range_regex.sub(repl_general_range, expr)
    
    # Pass 3: Substitute individual cell references
    def repl_single_cell(match):
        if match.group(1):
            return match.group(1)
        sheet_str = match.group('sheet')
        col_str = match.group('col')
        row_int = int(match.group('row'))
        lbl = resolve_cell_label(wb, sheet_name, sheet_str, col_str, row_int, formula_row)
        return f"\x01{lbl}\x02"
        
    expr = single_cell_regex.sub(repl_single_cell, expr)
    
    # Pass 4: Clean up operators and strip delimiters
    expr = expr.replace('*', ' x ')
    expr = expr.replace('\x01', '').replace('\x02', '')
    expr = re.sub(r'\s+', ' ', expr).strip()
    
    return f"{lhs_name} = {expr}"

print(f"Loading workbook: {excel_file}...")
try:
    wb = openpyxl.load_workbook(excel_file, data_only=False)
except Exception as e:
    print(f"Error loading workbook: {e}")
    sys.exit(1)

extracted_data = []

print("Iterating through sheets and translating formulas...")
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"Processing sheet: {sheet_name}...")
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                val_str = str(cell.value).strip()
                # Check if cell contains a formula
                if val_str.startswith("=") or getattr(cell, "data_type", "") == "f":
                    # Ignore simple single-cell references
                    if not simple_cell_ref_pattern.fullmatch(val_str):
                        word_formula = translate_formula_to_words(wb, sheet_name, cell, val_str)
                        extracted_data.append([sheet_name, cell.coordinate, val_str, word_formula])

print(f"Total actual formulas translated: {len(extracted_data)}")

print(f"Writing to {csv_file}...")
try:
    with open(csv_file, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Sheet", "Cell", "Formula", "Word Formula"])
        writer.writerows(extracted_data)
    print("CSV file generation complete.")
except Exception as e:
    print(f"Error writing CSV file: {e}")
    sys.exit(1)
