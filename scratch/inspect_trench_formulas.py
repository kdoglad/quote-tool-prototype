import openpyxl

excel_file = "XXkWp_Project Name_PCC_INTERNAL_V22.9 - DO NOT OPEN MAKE A COPY.xlsm"
wb = openpyxl.load_workbook(excel_file, data_only=False)
sheet = wb["3. Cabling"]

print("Trenching lookup table formulas in Column S (col 19):")
for r in range(6, 10):
    val_formula = sheet.cell(row=r, column=19).value
    val_data = openpyxl.load_workbook(excel_file, data_only=True)["3. Cabling"].cell(row=r, column=19).value
    print(f"Row {r} | Formula: {val_formula} | Evaluated: {val_data}")
