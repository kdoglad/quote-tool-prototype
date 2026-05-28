import openpyxl

excel_file = "XXkWp_Project Name_PCC_INTERNAL_V22.9 - DO NOT OPEN MAKE A COPY.xlsm"

wb_formula = openpyxl.load_workbook(excel_file, data_only=False)
sheet_formula = wb_formula["SOLAR"]

wb_data = openpyxl.load_workbook(excel_file, data_only=True)
sheet_data = wb_data["SOLAR"]

print("Row 39 values:")
for c in range(1, 12):
    col = openpyxl.utils.get_column_letter(c)
    val_formula = sheet_formula.cell(row=39, column=c).value
    val_data = sheet_data.cell(row=39, column=c).value
    print(f"Col {col} | Formula: {val_formula} | Data: {val_data}")
