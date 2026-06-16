import openpyxl

excel_file = "XXkWp_Project Name_PCC_INTERNAL_V22.9 - DO NOT OPEN MAKE A COPY.xlsm"

# Load without data_only to get formula text
wb_formula = openpyxl.load_workbook(excel_file, data_only=False)
sheet_formula = wb_formula["SOLAR"]

# Load with data_only to get evaluated values
wb_data = openpyxl.load_workbook(excel_file, data_only=True)
sheet_data = wb_data["SOLAR"]

print("Row | Component | Type | Size | Length | Formula | Value")
print("-" * 120)
for r in range(48, 53):
    comp = sheet_data.cell(row=r, column=3).value
    inclusion = sheet_data.cell(row=r, column=4).value
    size = sheet_data.cell(row=r, column=6).value
    length = sheet_data.cell(row=r, column=7).value
    
    formula_cell = sheet_formula.cell(row=r, column=9)
    val_cell = sheet_data.cell(row=r, column=9)
    
    formula_str = formula_cell.value
    if isinstance(formula_str, openpyxl.worksheet.formula.ArrayFormula):
        formula_str = f"{{ARRAY: {formula_str.text}}}"
        
    print(f"{r} | {comp} | {inclusion} | {size} | {length} | {formula_str} | {val_cell.value}")
