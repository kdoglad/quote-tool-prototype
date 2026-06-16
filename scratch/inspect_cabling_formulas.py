import openpyxl

excel_file = "XXkWp_Project Name_PCC_INTERNAL_V22.9 - DO NOT OPEN MAKE A COPY.xlsm"
wb = openpyxl.load_workbook(excel_file, data_only=True)  # Load with evaluated values first!

if "3. Cabling" in wb.sheetnames:
    sheet = wb["3. Cabling"]
    print("Sheet '3. Cabling' loaded successfully!")
    
    # Let's inspect the entire sheet content from columns A to L and rows 1 to 60
    for r in range(1, 60):
        row_vals = []
        for c in range(1, 15):
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {val}")
        if row_vals:
            print(" | ".join(row_vals))
else:
    print("Sheet '3. Cabling' not found in workbook sheet names:", wb.sheetnames)
