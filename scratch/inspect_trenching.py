import openpyxl

excel_file = "XXkWp_Project Name_PCC_INTERNAL_V22.9 - DO NOT OPEN MAKE A COPY.xlsm"
wb = openpyxl.load_workbook(excel_file, data_only=True)
sheet = wb["3. Cabling"]

print("Trenching lookup table:")
for r in range(6, 10):
    row_vals = []
    for c in range(17, 20):  # Q is 17, R is 18, S is 19
        col = openpyxl.utils.get_column_letter(c)
        val = sheet.cell(row=r, column=c).value
        row_vals.append(f"{col}{r}: {val}")
    print(" | ".join(row_vals))
