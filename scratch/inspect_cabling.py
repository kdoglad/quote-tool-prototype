import openpyxl

excel_file = "XXkWp_Project Name_PCC_INTERNAL_V22.9 - DO NOT OPEN MAKE A COPY.xlsm"
wb = openpyxl.load_workbook(excel_file, data_only=False)
sheet = wb["SOLAR"]

print("Row | Col A | Col B | Col C | Col D | Col E | Col F | Col G | Col H | Col I")
print("-" * 120)
for r in range(40, 55):
    row_vals = []
    for c in range(1, 11):
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        if val is not None:
            row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {val}")
    if row_vals:
        print(" | ".join(row_vals))
