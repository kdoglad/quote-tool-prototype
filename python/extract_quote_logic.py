import openpyxl
import json
import sys

excel_file = "XXkWp_Project Name_PCC_INTERNAL_V22.9 - DO NOT OPEN MAKE A COPY.xlsm"
output_file = "full_quote_logic.json"

def is_formula(cell):
    """Check if cell contains a formula"""
    if cell.value is None:
        return False
    val_str = str(cell.value).strip()
    return val_str.startswith("=") or getattr(cell, "data_type", "") == "f"

def is_manual_input(cell):
    """Check if cell is a manual input (has value but no formula)"""
    if cell.value is None:
        return False
    if is_formula(cell):
        return False
    # Check if it's a number or a meaningful string
    val = cell.value
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        val_stripped = val.strip()
        # Try to parse as number
        try:
            float(val_stripped)
            return True
        except ValueError:
            # Check if it's a meaningful input string (not just labels)
            if len(val_stripped) > 0 and not val_stripped.endswith(':'):
                return True
    return False

def get_cell_label(sheet, row, col):
    """Get label for a cell by checking left columns and above rows"""
    label_parts = []
    
    # Check left columns (1 and 2) for row label
    for c in range(min(col - 1, 2), 0, -1):
        val = sheet.cell(row=row, column=c).value
        if val and isinstance(val, str):
            val_clean = val.strip()
            if len(val_clean) > 1 and not val_clean.startswith('='):
                label_parts.append(val_clean)
                break
    
    # Check above rows for column header
    if row > 1:
        for r in range(row - 1, max(0, row - 20), -1):
            val = sheet.cell(row=r, column=col).value
            if val and isinstance(val, str):
                val_clean = val.strip()
                if len(val_clean) > 1 and not val_clean.startswith('='):
                    label_parts.append(val_clean)
                    break
    
    if label_parts:
        return " - ".join(label_parts)
    return None

print(f"Loading workbook: {excel_file}...")
try:
    wb = openpyxl.load_workbook(excel_file, data_only=False)
except Exception as e:
    print(f"Error loading workbook: {e}")
    sys.exit(1)

quote_logic = {
    "workbook": excel_file,
    "sheets": {}
}

print("Extracting logic from all visible sheets...")
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Skip hidden sheets
    if sheet.sheet_state == 'hidden':
        continue
    
    print(f"Processing sheet: {sheet_name}...")
    
    sheet_data = {
        "manual_inputs": [],
        "formulas": [],
        "calculations": []
    }
    
    # Iterate through all cells
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            
            cell_ref = cell.coordinate
            cell_row = cell.row
            cell_col = cell.column
            
            # Get label for context
            label = get_cell_label(sheet, cell_row, cell_col)
            
            if is_formula(cell):
                # Extract formula
                formula_str = str(cell.value).strip()
                
                # Get calculated value safely
                try:
                    calc_val = wb[sheet_name].cell(cell_row, cell_col).value
                    if calc_val is not None and not isinstance(calc_val, (str, int, float, bool)):
                        calc_val = str(calc_val)
                except:
                    calc_val = None
                
                formula_entry = {
                    "cell": cell_ref,
                    "row": cell_row,
                    "column": cell_col,
                    "label": label,
                    "formula": formula_str,
                    "calculated_value": calc_val
                }
                
                sheet_data["formulas"].append(formula_entry)
                
                # Also add to calculations with more detail
                calc_entry = {
                    "cell": cell_ref,
                    "label": label or cell_ref,
                    "formula": formula_str,
                    "type": "calculated"
                }
                sheet_data["calculations"].append(calc_entry)
                
            elif is_manual_input(cell):
                # Extract manual input
                input_entry = {
                    "cell": cell_ref,
                    "row": cell_row,
                    "column": cell_col,
                    "label": label,
                    "value": cell.value,
                    "type": type(cell.value).__name__
                }
                
                sheet_data["manual_inputs"].append(input_entry)
    
    # Only add sheet if it has data
    if sheet_data["manual_inputs"] or sheet_data["formulas"]:
        quote_logic["sheets"][sheet_name] = sheet_data
        print(f"  Found {len(sheet_data['manual_inputs'])} manual inputs and {len(sheet_data['formulas'])} formulas")

print(f"\nTotal sheets processed: {len(quote_logic['sheets'])}")
print(f"Writing to {output_file}...")

try:
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(quote_logic, f, indent=2, ensure_ascii=False)
    print(f"Successfully created {output_file}")
except Exception as e:
    print(f"Error writing JSON file: {e}")
    sys.exit(1)

print("\nExtraction complete!")
